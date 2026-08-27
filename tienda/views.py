import os
import json
import requests
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, get_object_or_404, redirect
from .models import Categoria, Producto, Pedido, ItemPedido, Cupon, BlogPost, ResenaProducto, ConfiguracionSitio
from .forms import CategoriaForm, ProductoForm, CuponForm, BlogPostForm, ResenaForm, ConfiguracionSitioForm
from django.contrib.auth.views import LoginView
from .carrito import Carrito
from django.contrib import messages
from django.core.mail import send_mail
import mercadopago
from django.conf import settings
from django.urls import reverse
from django.views.decorators.http import require_GET
from django.core.cache import cache
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
import re
from django.db import models

#----------------
#PAGINA PUBLICA
#----------------

def aplicar_ordenamiento(queryset, orden):
    if orden == 'precio_asc':
        return queryset.order_by('precio')
    elif orden == 'precio_desc':
        return queryset.order_by('-precio')
    elif orden == 'nombre':
        return queryset.order_by('nombre')
    elif orden == 'recientes':
        return queryset.order_by('-id')
    return queryset

def index(request):
    categorias = Categoria.objects.all()
    orden = request.GET.get('orden', '')
    productos = Producto.objects.filter(disponible=True)
    productos = aplicar_ordenamiento(productos, orden)
    productos_destacados = Producto.objects.filter(disponible=True).order_by('-id')[:4]
    resenas = ResenaProducto.objects.filter(aprobado=True).select_related('producto')[:8]

    return render(request, "index.html", {
        "categorias": categorias,
        "productos": productos,
        "productos_destacados": productos_destacados,
        "resenas": resenas,
        "orden_actual": orden
    })

def terminos_condiciones(request):
    return render(request, "terminos.html")


def categoria_detail(request, slug):
    categoria = get_object_or_404(Categoria, slug=slug)
    orden = request.GET.get('orden', '')

    productos = Producto.objects.filter(
        categoria=categoria,
        disponible=True
    )
    productos = aplicar_ordenamiento(productos, orden)

    return render(request, "categoria.html", {
        "categoria": categoria,
        "productos": productos,
        "orden_actual": orden
    })

@require_GET
def api_buscar_productos(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'productos': []})
    
    productos = Producto.objects.filter(
        models.Q(nombre__icontains=q) | models.Q(descripcion__icontains=q),
        disponible=True
    )[:6]
    
    data = []
    for p in productos:
        data.append({
            'id': p.id,
            'nombre': p.nombre,
            'precio': f"${p.precio:,}".replace(',', '.'),
            'url': p.get_absolute_url(),
            'imagen': p.imagen.url if p.imagen else ''
        })
    return JsonResponse({'productos': data})


def obtener_descuento_cupon(request, total_carrito):
    codigo = request.session.get('cupon_codigo')
    if not codigo:
        return None, 0
    try:
        cupon = Cupon.objects.get(codigo__iexact=codigo)
        valido, _ = cupon.es_valido()
        if valido:
            descuento = cupon.calcular_descuento(total_carrito)
            return cupon, descuento
    except Cupon.DoesNotExist:
        pass
    return None, 0


def aplicar_cupon(request):
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip().upper()
        if not codigo:
            messages.error(request, "Por favor ingresa un código de cupón.")
            return redirect('procesar_pedido')
        
        try:
            cupon = Cupon.objects.get(codigo__iexact=codigo)
            valido, msg = cupon.es_valido()
            if not valido:
                messages.error(request, msg)
            else:
                request.session['cupon_codigo'] = cupon.codigo
                messages.success(request, f"¡Cupón '{cupon.codigo}' aplicado exitosamente!")
        except Cupon.DoesNotExist:
            messages.error(request, "El código de cupón ingresado no existe.")
            
    return redirect('procesar_pedido')


def quitar_cupon(request):
    if 'cupon_codigo' in request.session:
        del request.session['cupon_codigo']
        messages.info(request, "Cupón removido.")
    return redirect('procesar_pedido')

def producto_detail(request, slug=None, id=None):
    if slug:
        # Si slug es numérico por accidente, intenta buscar por ID
        if slug.isdigit():
            producto = Producto.objects.filter(id=int(slug)).first()
            if producto:
                if producto.slug and producto.slug != slug:
                    return redirect(producto.get_absolute_url(), permanent=True)
                return render(request, "producto_detail.html", {"producto": producto})
        producto = get_object_or_404(Producto, slug=slug)
    elif id:
        producto = get_object_or_404(Producto, id=id)
        if producto.slug:
            return redirect(producto.get_absolute_url(), permanent=True)
    else:
        return redirect('index')

    resenas = producto.resenas.filter(aprobado=True)
    resena_form = ResenaForm()

    return render(request, "producto_detail.html", {
        "producto": producto,
        "resenas": resenas,
        "resena_form": resena_form,
    })

def producto_detail_by_id(request, id):
    producto = get_object_or_404(Producto, id=id)
    if producto.slug:
        return redirect(producto.get_absolute_url(), permanent=True)
    return render(request, "producto_detail.html", {
        "producto": producto
    })

# CARRITO

def agregar_al_carrito(request, producto_id):
    # 1. Inicializamos clase
    carrito = Carrito(request)
    producto = get_object_or_404(Producto, id=producto_id)

    try:
        cantidad = int(request.POST.get('cantidad', 1))
        if cantidad <= 0:
            messages.error(request, "La cantidad debe ser un número positivo.")
            return redirect(request.META.get('HTTP_REFERER', '/'))
    except ValueError:
        messages.error(request, "Cantidad no válida.")
        return redirect(request.META.get('HTTP_REFERER', '/'))
    
    # Opcional: si en el futuro agregas un input para elegir cantidad, lo lee aquí. Si no, usa 1.
    cantidad = int(request.POST.get('cantidad', 1) if request.method == 'POST' else 1)

    # 2. Verificamos si el producto tiene stock en absoluto (tu validación original)
    if producto.hay_stock():
        
        # 3. Intentamos agregar al carrito. Esto nos devolverá True (éxito) o False (pasó el límite)
        agregado_exitosamente = carrito.agregar(producto, cantidad)
        
        if agregado_exitosamente:
            messages.success(request, f'¡{producto.nombre} agregado a tu nido! 🪹')
        else:
            # Si devuelve False, es porque el cliente intentó agregar más de lo que tienes
            messages.warning(request, f'¡Límite alcanzado! Solo nos quedan {producto.stock} unidades de {producto.nombre} y ya están en tu nido. 🪹')
            
    else:
        # Si el stock en la base de datos es 0
        messages.error(request, f'Lo sentimos, {producto.nombre} está agotado por ahora.')
        
    # Redirigimos al catálogo o donde estaba el usuario y abrimos el carrito
    url_anterior = request.META.get('HTTP_REFERER', '/')
    if '?' in url_anterior:
        return redirect(url_anterior + '&cart=open')
    else:
        return redirect(url_anterior + '?cart=open')


def ver_carrito(request):
    
    return redirect('/?cart=open')


def procesar_pedido(request):
    carrito = Carrito(request)
    
    # Seguridad: Si el carro está vacío, no los dejamos pasar
    if len(carrito.carrito) == 0:
        messages.warning(request, "Tu nido está vacío. ¡Ve a pajarear al catálogo!")
        return redirect('index')
    
    for item_data in carrito:
        producto = item_data.get('producto_real')
        cantidad_pedida = item_data['cantidad']
        
        if not producto:
            messages.error(request, "Un producto de tu nido ya no está disponible.")
            return redirect('ver_carrito')
            
        if cantidad_pedida <= 0:
            messages.error(request, "Se detectó una cantidad inválida en tu nido. Por favor, revisa tus productos.")
            return redirect('ver_carrito')
            
        if cantidad_pedida > producto.stock:
            messages.warning(request, f"¡Atención! Mientras pensabas, el stock de '{producto.nombre}' bajó a {producto.stock} unidades. Por favor ajusta tu carrito.")
            return redirect('ver_carrito')

    total_bruto = carrito.get_total()
    cupon_obj, descuento_aplicado = obtener_descuento_cupon(request, total_bruto)
    total_final = max(0, total_bruto - descuento_aplicado)

    if request.method == 'POST':
        rut_ingresado = request.POST.get('rut', '')
        terminos_aceptados = request.POST.get('terminos_aceptados')

        if not terminos_aceptados:
            context = {
                'carrito': carrito,
                'cupon': cupon_obj,
                'descuento': descuento_aplicado,
                'total_final': total_final,
                'error_terminos': "Debes aceptar los Términos y Condiciones y Políticas de Devolución para realizar tu pedido.",
                'datos_previos': request.POST
            }
            return render(request, 'checkout.html', context)

        if not validar_rut_chileno(rut_ingresado):
            context = {
                'carrito': carrito,
                'cupon': cupon_obj,
                'descuento': descuento_aplicado,
                'total_final': total_final,
                'error_rut': "El RUT ingresado no es válido. Por favor, revísalo y escríbelo correctamente.",
                'datos_previos': request.POST
            }
            return render(request, 'checkout.html', context)

        # 1. Capturamos los datos del cliente desde el formulario
        pedido = Pedido.objects.create(
            nombre_completo=request.POST.get('nombre_completo'),
            rut=rut_ingresado,
            email=request.POST.get('email'),
            telefono=request.POST.get('telefono'),
            direccion=request.POST.get('direccion'),
            ciudad=request.POST.get('ciudad', 'Puerto Montt'),
            cupon=cupon_obj,
            descuento_aplicado=descuento_aplicado
        )
        
        if cupon_obj:
            cupon_obj.usos_actuales += 1
            cupon_obj.save()
            if 'cupon_codigo' in request.session:
                del request.session['cupon_codigo']

        # Guardamos el ID del pedido en la sesión para autorizar la vista de éxito
        request.session['pedido_autorizado'] = str(pedido.id)
        
        # 2. Guardamos los items del pedido
        for item in carrito:
            ItemPedido.objects.create(
                pedido=pedido,
                producto=item['producto_real'],
                precio=item['precio'],
                cantidad=item['cantidad']
            )

        # ========================================================
        # BLOQUE DE CORREO: ALERTA PARA ADMINISTRADORES (SILENCIOSO)
        # ========================================================
        if getattr(settings, 'EMAIL_HOST_USER', None):
            try:
                asunto_admin = f"🚨 NUEVO PEDIDO RARATIENDA#{pedido.codigo_orden} - {pedido.nombre_completo}"
                mensaje_admin = f"""¡Atención! Acaba de entrar un nuevo pedido.

Cliente: {pedido.nombre_completo}
Ciudad: {pedido.ciudad}
Total: ${total_final}
Teléfono: +56{pedido.telefono}

Revisa el panel de administración para ver el detalle completo.
www.raratienda.cl/panel
"""
                send_mail(
                    asunto_admin,
                    mensaje_admin,
                    settings.DEFAULT_FROM_EMAIL,
                    [settings.EMAIL_HOST_USER],
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Error silencioso al enviar alerta de pedido: {e}")

        # ========================================================
        # 🚀 INICIO INTEGRACIÓN MERCADO PAGO
        # ========================================================
        if not settings.MP_ACCESS_TOKEN:
            messages.info(request, "Entorno local: MERCADOPAGO_ACCESS_TOKEN no configurado en .env. Pedido registrado correctamente.")
            return redirect('pedido_confirmado', pedido_id=pedido.id)

        try:
            sdk = mercadopago.SDK(settings.MP_ACCESS_TOKEN)

            items_mp = []
            for item in carrito:
                items_mp.append({
                    "title": item['producto_real'].nombre,
                    "quantity": int(item['cantidad']),
                    "unit_price": int(item['precio']),
                    "currency_id": "CLP"
                })

            if descuento_aplicado > 0 and cupon_obj:
                items_mp.append({
                    "title": f"Descuento Cupón ({cupon_obj.codigo})",
                    "quantity": 1,
                    "unit_price": -int(descuento_aplicado),
                    "currency_id": "CLP"
                })

            url_exito = f"https://raratienda.cl/pedido-confirmado/{pedido.id}/"
            url_fallo = "https://raratienda.cl/?cart=open"
            url_webhook = "https://raratienda.cl/webhook/mercadopago/"

            preference_data = {
                "items": items_mp,
                "payer": {
                    "name": pedido.nombre_completo,
                    "email": pedido.email,
                },
                "back_urls": {
                    "success": url_exito,
                    "failure": url_fallo,
                    "pending": url_exito,
                },
                "auto_return": "approved",
                "external_reference": str(pedido.id),
                "notification_url": url_webhook,
            }

            preference_response = sdk.preference().create(preference_data)

            print("\n=== RESPUESTA DE MERCADO PAGO ===")
            print(preference_response)
            print("=================================\n")

            if "init_point" not in preference_response.get("response", {}):
                messages.error(request, "Hubo un problema al contactar a la pasarela de pago. Por favor intenta de nuevo.")
                return redirect('ver_carrito')
            
            init_point = preference_response["response"]["init_point"]
            return redirect(init_point)
        except Exception as e:
            print(f"Error al conectar con Mercado Pago: {e}")
            messages.error(request, f"Error con la pasarela de pago: {e}")
            return redirect('pedido_confirmado', pedido_id=pedido.id)

    return render(request, 'checkout.html', {
        'carrito': carrito,
        'cupon': cupon_obj,
        'descuento': descuento_aplicado,
        'total_final': total_final
    })



def restar_del_carrito(request, producto_id):
    carrito = Carrito(request)
    producto = get_object_or_404(Producto, id=producto_id)
    carrito.restar(producto)
    #return redirect('ver_carrito')
    url_anterior = request.META.get('HTTP_REFERER', '/')
    if '?' in url_anterior:
        return redirect(url_anterior + '&cart=open')
    else:
        return redirect(url_anterior + '?cart=open')

def quitar_del_carrito(request, producto_id):
    carrito = Carrito(request)
    producto = get_object_or_404(Producto, id=producto_id)
    carrito.eliminar(producto)
    #return redirect('ver_carrito')
    url_anterior = request.META.get('HTTP_REFERER', '/')
    if '?' in url_anterior:
        return redirect(url_anterior + '&cart=open')
    else:
        return redirect(url_anterior + '?cart=open')

def limpiar_carrito(request):
    carrito = Carrito(request)
    carrito.limpiar()
    return redirect('ver_carrito')


def pedido_confirmado(request, pedido_id):
    # --- 🛡️ VALIDACIÓN DE SEGURIDAD (Prevención de IDOR) ---
    pedido_autorizado = request.session.get('pedido_autorizado')
    if str(pedido_autorizado) != str(pedido_id):
        # Si el usuario intenta ver el pedido de otro, lo mandamos al inicio
        return redirect('index')
    # -------------------------------------------------------

    # Buscamos el pedido recién creado para mostrarle sus datos
    pedido = get_object_or_404(Pedido, id=pedido_id)
    carrito = Carrito(request)
    carrito.limpiar()
    return render(request, 'pedido_confirmado.html', {'pedido': pedido})


def validar_rut_chileno(rut):
    """Limpia y valida un RUT chileno usando el algoritmo de Módulo 11."""
    # 1. Quitar puntos, guiones y espacios, dejar en mayúscula
    rut_limpio = rut.replace(".", "").replace("-", "").replace(" ", "").upper()
    
    # 2. Validar que tenga el largo correcto y solo números + K
    if not re.match(r'^\d{7,8}[0-9K]$', rut_limpio):
        return False
        
    # 3. Separar cuerpo del dígito verificador
    cuerpo = rut_limpio[:-1]
    dv_ingresado = rut_limpio[-1]
    
    # 4. Cálculo matemático (Módulo 11)
    suma = 0
    multiplo = 2
    for c in reversed(cuerpo):
        suma += int(c) * multiplo
        multiplo += 1
        if multiplo == 8:
            multiplo = 2
            
    resto = suma % 11
    dv_esperado = 11 - resto
    
    if dv_esperado == 11:
        dv_esperado = "0"
    elif dv_esperado == 10:
        dv_esperado = "K"
    else:
        dv_esperado = str(dv_esperado)
        
    # 5. Comparar si el cálculo coincide con el del cliente
    return dv_ingresado == dv_esperado

#-----------------------
#PANEL DE ADMINISTRACION
#-----------------------

@staff_member_required(login_url='login')
def panel_productos(request):
    productos_list = Producto.objects.all().order_by('-id')

    # Capturamos los filtros actuales
    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")
    per_page = request.GET.get('per_page', '10') # Nuevo: cantidad por página

    # Aplicamos filtros
    if categoria_id:
        productos_list = productos_list.filter(categoria_id=categoria_id)

    if estado == "disponible":
        productos_list = productos_list.filter(disponible=True)
    elif estado == "agotado":
        productos_list = productos_list.filter(disponible=False)

    categorias = Categoria.objects.all()

    # Paginación usando la variable dinámica per_page
    paginator = Paginator(productos_list, int(per_page))
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "panel/productos.html", {
        "page_obj": page_obj,
        "categorias": categorias,
        # Devolvemos estas variables para que el HTML sepa qué hay filtrado
        "categoria_actual": categoria_id,
        "estado_actual": estado,
        "per_page": per_page
    })


@staff_member_required(login_url='login')
def crear_producto(request):
    # 1. Capturamos la URL de retorno
    next_url = request.GET.get('next') or request.POST.get('next') or 'panel_productos'
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Producto creado exitosamente.")
            # 2. Redirigimos al estado anterior
            return redirect(next_url)
    else:
        form = ProductoForm()

    return render(request, "panel/producto_form.html", {
        "form": form,
        "next": next_url  # 3. Lo mandamos al template
    })


@staff_member_required(login_url='login')
def crear_categoria(request):

    next_url = request.GET.get('next') or request.POST.get('next') or 'panel_productos'
    
    if request.method == 'POST':
        form = CategoriaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoria creada exitosamente.")
            return redirect(next_url)
    
    else:
        form = CategoriaForm()

    return render(request, "panel/categoria_form.html", {
        "form": form,
        "next": next_url
    })


@staff_member_required(login_url='login')
def toggle_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    # Cambiamos el estado (si es True pasa a False, y viceversa)
    producto.disponible = not producto.disponible
    producto.save()
    
    # Buscamos si hay una dirección de retorno
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    else:
        return redirect('panel_productos')


@staff_member_required(login_url='login')
def panel_home(request):
    from django.db import models
    from django.utils import timezone

    ahora = timezone.now()
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    pedidos_pagados = Pedido.objects.filter(pagado=True)
    ventas_mes = ItemPedido.objects.filter(pedido__pagado=True, pedido__creado__gte=inicio_mes).aggregate(
        total=models.Sum(models.F('precio') * models.F('cantidad'))
    )['total'] or 0

    top_productos = Producto.objects.filter(items_pedido__pedido__pagado=True).annotate(
        total_vendidos=models.Sum('items_pedido__cantidad')
    ).order_by('-total_vendidos')[:5]

    context = {
        'pedidos_pendientes': Pedido.objects.filter(pagado=False).count(),
        'pedidos_pagados': pedidos_pagados.count(),
        'productos_agotados': Producto.objects.filter(stock=0).count(),
        'ventas_mes': ventas_mes,
        'top_productos': top_productos,
        'ultimos_pedidos': Pedido.objects.all().order_by('-fecha_creacion')[:5],
    }
    return render(request, 'panel/dashboard.html', context)

class CustomLoginView(LoginView):
    template_name = "login.html"


@staff_member_required(login_url='login')
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    next_url = request.GET.get('next') or request.POST.get('next') or 'panel_productos'
    form = ProductoForm(request.POST or None, request.FILES or None, instance=producto)
    
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Producto '{producto.nombre}' actualizado.")
        return redirect(next_url)
        
    return render(request, "panel/producto_form.html", {
        "form": form,
        "editando": True,
        "next": next_url
    })

@staff_member_required(login_url='login')
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.delete()
    return redirect('panel_productos')

@staff_member_required(login_url='login')
def panel_pedidos(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')

    pedidos = Pedido.objects.all()

    if q:
        pedidos = pedidos.filter(
            models.Q(nombre_completo__icontains=q) |
            models.Q(rut__icontains=q) |
            models.Q(email__icontains=q) |
            models.Q(id__icontains=q.replace('#', ''))
        )

    if estado:
        pedidos = pedidos.filter(estado=estado)

    pedidos = pedidos.order_by('-id')

    return render(request, 'panel/pedidos.html', {
        'pedidos': pedidos,
        'q': q,
        'estado_filtro': estado,
        'estados_choices': Pedido.ESTADO_CHOICES
    })

@staff_member_required(login_url='login')
def detalle_pedido(request, id):
    pedido = get_object_or_404(Pedido.objects.prefetch_related('items__producto'), id=id)
    return render(request, 'panel/detalle_pedido.html', {'pedido': pedido, 'estados_choices': Pedido.ESTADO_CHOICES})


@staff_member_required(login_url='login')
def actualizar_estado_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        empresa = request.POST.get('empresa_transporte', '').strip()
        seguimiento = request.POST.get('numero_seguimiento', '').strip()

        if nuevo_estado in dict(Pedido.ESTADO_CHOICES):
            pedido.estado = nuevo_estado
        pedido.empresa_transporte = empresa
        pedido.numero_seguimiento = seguimiento
        
        if nuevo_estado == 'PAGADO' and not pedido.pagado:
            pedido.confirmar_pago()
        else:
            pedido.save()

        messages.success(request, f"Estado del pedido #{pedido.codigo_orden} actualizado a '{pedido.get_estado_display()}'.")
    return redirect('detalle_pedido', id=id)


@staff_member_required(login_url='login')
def enviar_seguimiento_email(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if not pedido.numero_seguimiento:
        messages.error(request, "Debes ingresar un número de seguimiento primero.")
        return redirect('detalle_pedido', id=id)

    try:
        asunto = f"📦 Tu pedido #{pedido.codigo_orden} de Rara Tienda ya va en camino"
        mensaje = f"""Hola {pedido.nombre_completo},

¡Te tenemos excelentes noticias! Tu pedido #{pedido.codigo_orden} ha sido enviado.

Detalles del despacho:
- Transporte: {pedido.empresa_transporte or 'Empresa de Envíos'}
- Código / Nº de Seguimiento: {pedido.numero_seguimiento}

Puedes realizar el seguimiento de tu paquete directamente en el sitio web del transporte.

¡Muchas gracias por comprar en Rara Tienda! 🦉
"""
        send_mail(
            asunto,
            mensaje,
            settings.DEFAULT_FROM_EMAIL,
            [pedido.email],
            fail_silently=False
        )
        messages.success(request, f"Correo con número de seguimiento enviado a {pedido.email}.")
    except Exception as e:
        messages.error(request, f"No se pudo enviar el correo: {e}")

    return redirect('detalle_pedido', id=id)


@staff_member_required(login_url='login')
def confirmar_pago_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if request.method == 'POST':
        with transaction.atomic():
            pedido = Pedido.objects.select_for_update().get(id=id)
            if pedido.pagado:
                messages.info(request, f'El pedido #{pedido.codigo_orden} ya estaba pagado.')
                return redirect('detalle_pedido', id=pedido.id)
            pedido.confirmar_pago()
            pedido.estado = 'PAGADO'
            pedido.save()
        
        asunto = f'¡Pago Confirmado! Pedido #{pedido.codigo_orden} en Rara Tienda 🦉'
        mensaje = f'''Hola {pedido.nombre_completo},

¡Tenemos excelentes noticias! Hemos confirmado el pago de tu pedido.

Tu nido de productos ya está siendo preparado con mucho cariño para ser enviado a {pedido.direccion}, {pedido.ciudad}. 

Te avisaremos por esta misma vía en cuanto el paquete inicie su vuelo.

¡Gracias por apoyar el arte y la naturaleza!
El equipo de Rara Tienda.
'''
        try:
            send_mail(
                asunto, 
                mensaje, 
                settings.DEFAULT_FROM_EMAIL, 
                [pedido.email], 
                fail_silently=False
            )
            messages.success(request, f'Pago confirmado, stock actualizado y correo enviado a {pedido.email}.')
        except Exception as e:
            messages.warning(request, f'Pago confirmado y stock descontado, pero no se pudo enviar el correo automático.')
            print(f"Error SMTP: {e}")
            
    return redirect('detalle_pedido', id=pedido.id)


@staff_member_required(login_url='login')
def panel_cupones(request):
    cupones = Cupon.objects.all().order_by('-id')
    return render(request, 'panel/cupones.html', {'cupones': cupones})


@staff_member_required(login_url='login')
def crear_cupon(request):
    form = CuponForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Cupón creado exitosamente.")
        return redirect('panel_cupones')
    return render(request, 'panel/cupon_form.html', {'form': form, 'editando': False})


@staff_member_required(login_url='login')
def editar_cupon(request, id):
    cupon = get_object_or_404(Cupon, id=id)
    form = CuponForm(request.POST or None, instance=cupon)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Cupón '{cupon.codigo}' actualizado exitosamente.")
        return redirect('panel_cupones')
    return render(request, 'panel/cupon_form.html', {'form': form, 'editando': True})


@staff_member_required(login_url='login')
def toggle_cupon(request, id):
    cupon = get_object_or_404(Cupon, id=id)
    cupon.activo = not cupon.activo
    cupon.save()
    messages.success(request, f"Estado del cupón '{cupon.codigo}' cambiado.")
    return redirect('panel_cupones')


@staff_member_required(login_url='login')
def exportar_pedidos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Rara Tienda"

    fill_header = PatternFill(start_color="FCA311", end_color="FCA311", fill_type="solid")
    font_header = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(left=Side(style='thin', color='DDDDDD'),
                         right=Side(style='thin', color='DDDDDD'),
                         top=Side(style='thin', color='DDDDDD'),
                         bottom=Side(style='thin', color='DDDDDD'))

    headers = ['ID Orden', 'Cliente', 'RUT', 'Email', 'Teléfono', 'Ciudad', 'Dirección', 'Estado', 'Pagado', 'Total', 'Transporte', 'Nº Seguimiento', 'Fecha']
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 18

    pedidos = Pedido.objects.all().order_by('-id')
    
    for row_num, p in enumerate(pedidos, start=2):
        ws.append([
            f"#{p.codigo_orden}",
            p.nombre_completo,
            p.rut,
            p.email,
            f"+56{p.telefono}",
            p.ciudad,
            p.direccion,
            p.get_estado_display(),
            "Sí" if p.pagado else "No",
            f"${p.get_total_cost() - p.descuento_aplicado}",
            p.empresa_transporte,
            p.numero_seguimiento,
            p.creado.strftime("%Y-%m-%d %H:%M")
        ])
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [1, 3, 5, 8, 9, 10, 13]:
                cell.alignment = align_center

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Pedidos_Rara_Tienda.xlsx"'
    wb.save(response)
    return response


@staff_member_required(login_url='login')
def exportar_stock_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Rara Tienda"

    fill_header = PatternFill(start_color="FCA311", end_color="FCA311", fill_type="solid")
    font_header = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(left=Side(style='thin', color='DDDDDD'),
                         right=Side(style='thin', color='DDDDDD'),
                         top=Side(style='thin', color='DDDDDD'),
                         bottom=Side(style='thin', color='DDDDDD'))

    font_agotado = Font(color="E74C3C", bold=True)
    font_disponible = Font(color="27AE60", bold=True)

    headers = ['ID', 'Producto', 'Stock Actual', 'Precio', 'Estado']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

    productos = Producto.objects.all().order_by('nombre')
    
    for row_num, p in enumerate(productos, start=2): 
        estado = "Disponible" if p.stock > 0 else "Agotado"
        ws.append([p.id, p.nombre, p.stock, f"${p.precio}", estado])
        
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [1, 3, 4, 5]: 
                cell.alignment = align_center
            if col_num == 5:
                cell.font = font_agotado if estado == "Agotado" else font_disponible

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventario_Rara_Tienda.xlsx"'
    wb.save(response)
    return response

#---------------------
# RADAR BIG DAY
#---------------------
def generar_avistamientos_fallback(ebird_path):
    """Genera avistamientos de respaldo realistas si la API Key de eBird no está configurada o falla"""
    partes = ebird_path.split('/')
    region = 'CL-RM'
    for p in partes:
        if p.startswith('CL-'):
            region = p
            break
            
    regiones_map = {
        'CL-AP': (-18.4783, -70.3126, ['Desembocadura del Río Lluta', 'Valle de Azapa', 'Putre']),
        'CL-TA': (-20.2133, -70.1503, ['Salar de Huasco', 'Iquique Costanera', 'Pica']),
        'CL-AN': (-23.6509, -70.3975, ['La Portada', 'Humedal Aguada La Chimba', 'Mejillones']),
        'CL-AT': (-27.3668, -70.3323, ['Bahía Inglesa', 'Parque Nacional Pan de Azúcar', 'Huasco']),
        'CL-CO': (-29.9533, -71.3395, ['Humedal El Culebrón', 'Punta de Choros', 'Reserva Chinchillas']),
        'CL-VS': (-33.0472, -71.6127, ['Humedal El Peral', 'Concón Desembocadura', 'Laguna Verde']),
        'CL-RM': (-33.4489, -70.6693, ['Parque Bicentenario', 'Cajón del Maipo', 'Humedal Batuco', 'Cerro San Cristóbal']),
        'CL-LI': (-34.1701, -70.7407, ['Reserva Nacional Río de los Cipreses', 'Pichilemu', 'Rapel']),
        'CL-ML': (-35.4264, -71.6554, ['Reserva Altos de Lircay', 'Putú', 'Vilches']),
        'CL-NB': (-36.6066, -72.1034, ['Cobquecura', 'Nevados de Chillán', 'Quillón']),
        'CL-BI': (-36.8270, -73.0498, ['Desembocadura del Bío Bío', 'Dichato', 'Lenga']),
        'CL-AR': (-38.7359, -72.5904, ['Lago Villarrica', 'Parque Conguillío', 'Puerto Saavedra']),
        'CL-LR': (-39.8142, -73.2459, ['Santuario Carlos Anwandter', 'Niebla', 'Lago Ranco']),
        'CL-LL': (-41.4693, -72.9424, ['Seno de Reloncaví', 'Chiloé Castro', 'Frutillar', 'Alerce Andino']),
        'CL-AI': (-45.5752, -72.0662, ['Coyhaique Alto', 'Puerto Aysén', 'Lago General Carrera']),
        'CL-MA': (-53.1638, -70.9171, ['Estrecho de Magallanes', 'Torres del Paine', 'Puerto Natales']),
    }
    
    base_lat, base_lng, lugares = regiones_map.get(region, (-33.4489, -70.6693, ['Reserva Natural', 'Costanera', 'Parque Central']))
    
    especies_muestra = [
        ("austhr1", "Zorzal patagónico", "Turdus falcklandii"),
        ("rucspa1", "Chincol", "Zonotrichia capensis"),
        ("grbfir1", "Picaflor chico", "Sephanoides sephaniodes"),
        ("tuttyr1", "Cachudito común", "Anairetes parulus"),
        ("houwre4", "Chercán común", "Troglodytes musculus"),
        ("eardov1", "Tórtola", "Zenaida auriculata"),
        ("rocpig", "Paloma doméstica", "Columba livia"),
        ("ameoys", "Pilpilén común", "Haematopus palliatus"),
        ("lesrhe2", "Suri/Ñandú", "Rhea pennata"),
        ("bkbplo", "Chorlo ártico", "Pluvialis squatarola"),
        ("blhher1", "Garza cuca", "Ardea cocoi"),
        ("blnhea1", "Huairavo", "Nycticorax nycticorax"),
        ("chihaw1", "Peuco", "Parabuteo unicinctus"),
        ("bkycar1", "Tiuque", "Milvago chimango"),
    ]
    
    fallback_data = []
    import random
    from django.utils import timezone
    ahora_str = timezone.now().strftime("%Y-%m-%d %H:%M")
    
    for idx, (code, com, sci) in enumerate(especies_muestra):
        num_obs = random.randint(1, 3)
        for o in range(num_obs):
            offset_lat = (random.random() - 0.5) * 0.15
            offset_lng = (random.random() - 0.5) * 0.15
            lugar = lugares[o % len(lugares)]
            
            fallback_data.append({
                "speciesCode": code,
                "comName": com,
                "sciName": sci,
                "locId": f"L{idx}{o}",
                "locName": lugar,
                "obsDt": ahora_str,
                "howMany": random.randint(1, 6),
                "lat": round(base_lat + offset_lat, 4),
                "lng": round(base_lng + offset_lng, 4),
                "obsValid": True,
                "obsReviewed": False,
                "locationPrivate": False,
                "subId": f"S{idx}{o}"
            })
            
    return fallback_data

@require_GET
def ebird_proxy(request, ebird_path):
    """Proxy para eBird que inyecta la API Key en secreto, bloquea robos de código y cachea respuestas"""
    cache_key = f"ebird_{ebird_path}_{request.GET.urlencode()}"
    
    datos_cacheados = cache.get(cache_key)
    if datos_cacheados:
        return JsonResponse(datos_cacheados, safe=False)

    api_key = getattr(settings, 'EBIRD_API_KEY', None) or os.environ.get('EBIRD_API_KEY')
    
    if api_key:
        url = f"https://api.ebird.org/v2/{ebird_path}"
        headers = {"X-eBirdApiToken": api_key}
        params = request.GET.dict()
        
        try:
            response_ebird = requests.get(url, headers=headers, params=params)
            response_ebird.raise_for_status()
            
            datos_nuevos = response_ebird.json()
            cache.set(cache_key, datos_nuevos, 300)
            return JsonResponse(datos_nuevos, safe=False)
        except requests.RequestException as e:
            print(f"⚠️ Error al conectar con eBird API: {e}. Usando avistamientos de respaldo.")
    
    # Fallback automático si no hay API Key o falla eBird
    datos_fallback = generar_avistamientos_fallback(ebird_path)
    cache.set(cache_key, datos_fallback, 300)
    return JsonResponse(datos_fallback, safe=False)
    

def get_species_dict(request):
    """Lee el diccionario local y lo devuelve como JSON puro con CORS público"""
    file_path = os.path.join(settings.BASE_DIR, 'tienda', 'species.json')
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # 1. Guardamos la respuesta en una variable en vez de retornarla de inmediato
        response = JsonResponse(data, safe=False)
        
        # 2. Le pegamos la cabecera CORS universal
        response["Access-Control-Allow-Origin"] = "*"
        
        # 3. Ahora sí, devolvemos la respuesta
        return response
        
    except FileNotFoundError:
        return JsonResponse({"error": "Archivo no encontrado"}, status=404)
    
@csrf_exempt
def webhook_mercadopago(request):
    """
    Esta es la ruta secreta que Mercado Pago visitará por detrás 
    cuando un cliente pague con éxito.
    """
    if request.method == 'POST':
        try:
            # 1. Leemos el mensaje en formato JSON que envía MP
            data = json.loads(request.body)
            
            # 2. Verificamos si nos están avisando de un "pago"
            if data.get("type") == "payment" or data.get("action") == "payment.created":
                # Capturamos el ID del pago
                payment_id = data.get("data", {}).get("id")
                
                if payment_id:
                    # 3. Consultamos directamente a MP para confirmar que no sea un aviso falso (Seguridad)
                    sdk = mercadopago.SDK(settings.MP_ACCESS_TOKEN)
                    payment_info = sdk.payment().get(payment_id)
                    payment = payment_info.get("response")
                    
                    # 4. Si MP nos confirma que el pago está APROBADO 🟢
                    if payment and payment.get("status") == "approved":
                        # Rescatamos el ID de tu pedido (el que enviamos al crear la preferencia)
                        pedido_id = payment.get("external_reference")
                        
                        if pedido_id:
                            pago_procesado = False
                            pedido = None
                            with transaction.atomic():
                                # 5. Buscamos el pedido en tu base de datos y bloqueamos fila para evitar doble proceso
                                pedido = Pedido.objects.select_for_update().filter(id=pedido_id).first()
                                # Si el pedido existe y aún no estaba pagado...
                                if pedido and not pedido.pagado:
                                    # ¡MAGIA! Ejecutamos tu súper función que descuenta stock
                                    pedido.confirmar_pago()
                                    
                                    # Guardamos el ID de transacción de MP por si hay devoluciones a futuro
                                    pedido.id_transaccion = str(payment_id)
                                    pedido.save()
                                    pago_procesado = True
                            
                            if pago_procesado and pedido:
                                print(f"✅ ¡ÉXITO! Pedido #{pedido.id} pagado y stock descontado.")

                                asunto = f'¡Pago Recibido! Tu pedido #{pedido.codigo_orden} está en camino 🦉'
                                mensaje = f'''¡Hola {pedido.nombre_completo}!

Te escribimos de Rara Tienda para contarte que hemos recibido el pago de tu pedido #{pedido.codigo_orden} con éxito a través de Mercado Pago. ✨

¿Qué viene ahora?
Estamos preparando tu paquete con mucho cuidado para que llegue perfecto a tu casa. 

📍 Destino: {pedido.direccion}, {pedido.ciudad}.

En cuanto realicemos el envío, te contactaremos por WhatsApp al +56{pedido.telefono} para enviarte el comprobante y el número de seguimiento.

¡Gracias por confiar en Rara Tienda y apoyar el arte local!

Un gran abrazo,
El equipo de Rara Tienda.
www.raratienda.cl
'''
                                try:
                                    send_mail(
                                        asunto,
                                        mensaje,
                                        settings.DEFAULT_FROM_EMAIL,
                                        [pedido.email],
                                        fail_silently=False,
                                    )
                                    print(f"✅ Webhook: Pago procesado y correo enviado a {pedido.email}")
                                except Exception as mail_error:
                                    # Si el correo falla, igual el pago ya quedó registrado
                                    print(f"⚠️ Webhook: Pago OK pero falló el correo: {mail_error}")
                            elif pedido and pedido.pagado:
                                print(f"ℹ️ Webhook duplicado ignorado para pedido #{pedido.id}.")

            # SIEMPRE debemos responder 200 OK, sino MP pensará que falló y enviará el aviso de nuevo
            return JsonResponse({"status": "ok"}, status=200)

        except Exception as e:
            print(f"❌ Error en Webhook: {e}")
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    # Si alguien intenta entrar por la URL normal (GET), lo rechazamos
    return JsonResponse({"status": "method not allowed"}, status=405)

#----------------
# JUEGOS
#----------------

def menu_juegos(request):
    return render(request, "juegos/menu_juegos.html")

def quien_es_esta_ave(request):
    return render(request, "juegos/quien_es_esta_ave.html")

def memorice(request):
    return render(request, "juegos/memorice.html")

def letras_locas(request):
    return render(request, "juegos/letras_locas.html")

def minijuego(request):
    return render(request, "juegos/minijuego.html")

def rosco(request):
    return render(request, "juegos/rosco.html")

def aves_en_tu_zona(request):
    return render(request, "juegos/aves_en_tu_zona.html")

def radar_realtime(request):
    return render(request, "juegos/radar_realtime.html")


#--------------------------------------
# BLOG, RESEÑAS Y CONFIGURACIÓN SITIO
#--------------------------------------

def agregar_resena(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        form = ResenaForm(request.POST)
        if form.is_valid():
            resena = form.save(commit=False)
            resena.producto = producto
            resena.save()
            messages.success(request, "¡Gracias por tu reseña! Ha sido publicada.")
        else:
            messages.error(request, "Por favor revisa los campos ingresados en tu reseña.")
    return redirect(producto.get_absolute_url())

def blog_list(request):
    posts_list = BlogPost.objects.filter(publicado=True)
    paginator = Paginator(posts_list, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, "blog/blog_list.html", {"page_obj": page_obj})

def blog_detail(request, slug):
    post = get_object_or_404(BlogPost, slug=slug, publicado=True)
    otros_posts = BlogPost.objects.filter(publicado=True).exclude(id=post.id)[:3]
    productos_destacados = Producto.objects.filter(disponible=True).order_by('?')[:4]
    return render(request, "blog/blog_detail.html", {
        "post": post,
        "otros_posts": otros_posts,
        "productos_destacados": productos_destacados,
    })

def panel_configuracion(request):
    if not request.user.is_staff:
        return redirect('login')
    config = ConfiguracionSitio.get_solo()
    if request.method == 'POST':
        form = ConfiguracionSitioForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración del sitio actualizada correctamente.")
            return redirect('panel_configuracion')
    else:
        form = ConfiguracionSitioForm(instance=config)
    return render(request, "panel/configuracion.html", {"form": form, "config": config})

def panel_blog(request):
    if not request.user.is_staff:
        return redirect('login')
    posts = BlogPost.objects.all()
    return render(request, "panel/blog_list.html", {"posts": posts})

def crear_blog_post(request):
    if not request.user.is_staff:
        return redirect('login')
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Artículo publicado exitosamente.")
            return redirect('panel_blog')
    else:
        form = BlogPostForm()
    return render(request, "panel/blog_form.html", {"form": form, "titulo_pagina": "Nuevo Artículo del Blog"})

def editar_blog_post(request, id):
    if not request.user.is_staff:
        return redirect('login')
    post = get_object_or_404(BlogPost, id=id)
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            form.save()
            messages.success(request, "Artículo actualizado correctamente.")
            return redirect('panel_blog')
    else:
        form = BlogPostForm(instance=post)
    return render(request, "panel/blog_form.html", {"form": form, "post": post, "titulo_pagina": "Editar Artículo"})

def eliminar_blog_post(request, id):
    if not request.user.is_staff:
        return redirect('login')
    post = get_object_or_404(BlogPost, id=id)
    if request.method == 'POST':
        post.delete()
        messages.success(request, "Artículo eliminado.")
    return redirect('panel_blog')

def panel_resenas(request):
    if not request.user.is_staff:
        return redirect('login')
    resenas = ResenaProducto.objects.select_related('producto').all()
    return render(request, "panel/resenas.html", {"resenas": resenas})

def toggle_resena(request, id):
    if not request.user.is_staff:
        return redirect('login')
    resena = get_object_or_404(ResenaProducto, id=id)
    resena.aprobado = not resena.aprobado
    resena.save()
    estado = "aprobada" if resena.aprobado else "ocultada"
    messages.info(request, f"Reseña {estado}.")
    return redirect('panel_resenas')

def eliminar_resena(request, id):
    if not request.user.is_staff:
        return redirect('login')
    resena = get_object_or_404(ResenaProducto, id=id)
    if request.method == 'POST':
        resena.delete()
        messages.success(request, "Reseña eliminada.")
    return redirect('panel_resenas')


def evaluar_compra(request, token):
    pedido = get_object_or_404(Pedido, token_resena=token)
    items = pedido.items.select_related('producto').all()
    
    if request.method == 'POST':
        resenas_creadas = 0
        for item in items:
            prod_id = item.producto.id
            calif_key = f"calificacion_{prod_id}"
            coment_key = f"comentario_{prod_id}"
            
            if calif_key in request.POST:
                try:
                    calificacion = int(request.POST.get(calif_key, 5))
                except ValueError:
                    calificacion = 5
                comentario = request.POST.get(coment_key, '').strip()
                
                if comentario:
                    ResenaProducto.objects.create(
                        producto=item.producto,
                        pedido=pedido,
                        nombre_cliente=pedido.nombre_completo,
                        email_cliente=pedido.email,
                        calificacion=calificacion,
                        comentario=comentario,
                        comprador_verificado=True,
                        aprobado=True
                    )
                    resenas_creadas += 1

        if resenas_creadas > 0:
            return render(request, "evaluaciones/evaluar_exito.html", {"pedido": pedido, "resenas_creadas": resenas_creadas})
        else:
            messages.warning(request, "Por favor escribe al menos un comentario en tu calificación.")

    return render(request, "evaluaciones/evaluar_compra.html", {"pedido": pedido, "items": items})


def enviar_resena_email(request, id):
    if not request.user.is_staff:
        return redirect('login')
    pedido = get_object_or_404(Pedido, id=id)
    url_resena = pedido.get_enlace_resena()
    
    asunto = f"¡Cuéntanos tu experiencia con el Pedido #{pedido.codigo_orden}! 🌟 — Rara Tienda"
    mensaje = f"""Hola {pedido.nombre_completo},

¡Esperamos que estés disfrutando tus productos de Rara Tienda!

Nos encantaría saber tu opinión sobre tu compra. Tu valoración ayuda a otros amantes de la naturaleza a conocer nuestros productos:

👉 Evalúa tus productos aquí: {url_resena}

¡Muchas gracias por apoyar nuestra tienda!

Con cariño,
El equipo de Rara Tienda
https://www.raratienda.cl/
"""
    try:
        send_mail(
            asunto,
            mensaje,
            settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@raratienda.cl',
            [pedido.email],
            fail_silently=False
        )
        messages.success(request, f"Correo de reseña enviado con éxito a {pedido.email}.")
    except Exception as e:
        messages.error(request, f"Error al enviar el correo: {e}")
        
    return redirect('detalle_pedido', id=pedido.id)


def evaluar_producto_directo(request, id):
    producto = get_object_or_404(Producto, id=id)
    
    if request.method == 'POST':
        nombre_cliente = request.POST.get('nombre_cliente', '').strip()
        email_cliente = request.POST.get('email_cliente', '').strip()
        try:
            calificacion = int(request.POST.get('calificacion', 5))
        except ValueError:
            calificacion = 5
        comentario = request.POST.get('comentario', '').strip()
        
        if nombre_cliente and comentario:
            ResenaProducto.objects.create(
                producto=producto,
                nombre_cliente=nombre_cliente,
                email_cliente=email_cliente,
                calificacion=calificacion,
                comentario=comentario,
                comprador_verificado=True,
                aprobado=True
            )
            return render(request, "evaluaciones/evaluar_producto_exito.html", {"producto": producto, "nombre_cliente": nombre_cliente})
        else:
            messages.warning(request, "Por favor completa tu nombre y comentario.")

    return render(request, "evaluaciones/evaluar_producto_directo.html", {"producto": producto})


@staff_member_required(login_url='login')
def panel_guia(request):
    return render(request, "panel/guia.html", {"titulo_pagina": "Guía y Ayuda del Panel"})