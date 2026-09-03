import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.views import LoginView
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse
from django.db import models, transaction
from django.utils import timezone

from ..models import Categoria, Producto, Pedido, ItemPedido, Cupon, BlogPost, ResenaProducto, ConfiguracionSitio, LogProducto, LogPedido, MetricaDiaria, MetricaProducto
from ..forms import CategoriaForm, ProductoForm, CuponForm, BlogPostForm, ConfiguracionSitioForm


class CustomLoginView(LoginView):
    template_name = "login.html"


@staff_member_required(login_url='login')
def panel_home(request):
    ahora = timezone.now()
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    pedidos_pagados = Pedido.objects.filter(pagado=True)
    # Suma de items
    suma_items = ItemPedido.objects.filter(pedido__pagado=True, pedido__creado__gte=inicio_mes).aggregate(
        total=models.Sum(models.F('precio') * models.F('cantidad'))
    )['total'] or 0

    # Suma de descuentos de esos mismos pedidos
    suma_descuentos = Pedido.objects.filter(pagado=True, creado__gte=inicio_mes).aggregate(
        total=models.Sum('descuento_aplicado')
    )['total'] or 0

    ventas_mes = max(0, suma_items - suma_descuentos)

    top_productos = Producto.objects.filter(items_pedido__pedido__pagado=True).annotate(
        total_vendidos=models.Sum('items_pedido__cantidad')
    ).order_by('-total_vendidos')[:5]
    
    hoy = timezone.localdate()
    metrica_hoy = MetricaDiaria.objects.filter(fecha=hoy).first()
    visitas_humanos_hoy = metrica_hoy.visitas_humanos if metrica_hoy else 0
    visitas_bots_hoy = metrica_hoy.visitas_bots if metrica_hoy else 0
    
    inicio_mes_date = inicio_mes.date()
    visitas_mes = MetricaDiaria.objects.filter(fecha__gte=inicio_mes_date).aggregate(
        humanos=models.Sum('visitas_humanos'),
        bots=models.Sum('visitas_bots')
    )
    humanos_mes = visitas_mes['humanos'] or 0
    bots_mes = visitas_mes['bots'] or 0
    
    top_vistos = Producto.objects.annotate(
        total_vistas=models.Sum('metricas__vistas')
    ).exclude(total_vistas=None).order_by('-total_vistas')[:5]

    context = {
        'pedidos_pendientes': Pedido.objects.filter(pagado=False).count(),
        'pedidos_pagados': pedidos_pagados.count(),
        'productos_agotados': Producto.objects.filter(stock=0).count(),
        'ventas_mes': ventas_mes,
        'top_productos': top_productos,
        'top_vistos': top_vistos,
        'visitas_humanos_hoy': visitas_humanos_hoy,
        'visitas_bots_hoy': visitas_bots_hoy,
        'humanos_mes': humanos_mes,
        'bots_mes': bots_mes,
        'ultimos_pedidos': Pedido.objects.all().order_by('-fecha_creacion')[:5],
    }
    return render(request, 'panel/dashboard.html', context)


@staff_member_required(login_url='login')
def panel_productos(request):
    productos_list = Producto.objects.all().order_by('-id')

    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")
    per_page = request.GET.get('per_page', '10')

    if categoria_id:
        productos_list = productos_list.filter(categoria_id=categoria_id)

    if estado == "disponible":
        productos_list = productos_list.filter(disponible=True)
    elif estado == "agotado":
        productos_list = productos_list.filter(disponible=False)

    categorias = Categoria.objects.all()

    paginator = Paginator(productos_list, int(per_page))
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "panel/productos.html", {
        "page_obj": page_obj,
        "categorias": categorias,
        "categoria_actual": categoria_id,
        "estado_actual": estado,
        "per_page": per_page
    })


@staff_member_required(login_url='login')
def crear_producto(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'panel_productos'
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()

            # Guardar imágenes adicionales de galería si se enviaron
            imagenes_extra = request.FILES.getlist('imagenes_adicionales')
            for idx, file in enumerate(imagenes_extra):
                ImagenProducto.objects.create(
                    producto=producto,
                    imagen=file,
                    orden=idx
                )

            usuario_log = request.user if request.user.is_authenticated else None
            detalles_str = f"Precio inicial: ${producto.precio} | Stock inicial: {producto.stock} | Categoría: {producto.categoria.nombre if producto.categoria else 'Sin categoría'}"
            if imagenes_extra:
                detalles_str += f" | {len(imagenes_extra)} imágenes adicionales subidas"
            LogProducto.objects.create(
                producto_id=producto.id,
                nombre_producto=producto.nombre,
                accion='CREACION',
                usuario=usuario_log,
                detalles=detalles_str
            )
            messages.success(request, f"Producto '{producto.nombre}' creado exitosamente.")
            return redirect(next_url)
    else:
        form = ProductoForm()

    return render(request, "panel/producto_form.html", {
        "form": form,
        "next": next_url
    })


@staff_member_required(login_url='login')
def crear_categoria(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'panel_productos'
    
    if request.method == 'POST':
        form = CategoriaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoria creada exitosamente.")
            return redirect(next_url)
    else:
        form = CategoriaForm()

    return render(request, "panel/categoria_form.html", {
        "form": form,
        "next": next_url
    })


@staff_member_required(login_url='login')
def toggle_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.disponible = not producto.disponible
    producto.save()
    
    usuario_log = request.user if request.user.is_authenticated else None
    estado_str = "Activado en catálogo (Visible)" if producto.disponible else "Desactivado de catálogo (Oculto)"
    LogProducto.objects.create(
        producto_id=producto.id,
        nombre_producto=producto.nombre,
        accion='TOGGLE',
        usuario=usuario_log,
        detalles=f"Estado del producto cambiado a: {estado_str}"
    )
    
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    else:
        return redirect('panel_productos')


@staff_member_required(login_url='login')
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    nombre_ant = producto.nombre
    precio_ant = producto.precio
    stock_ant = producto.stock
    dispon_ant = producto.disponible
    cat_ant = producto.categoria.nombre if producto.categoria else "Sin categoría"

    next_url = request.GET.get('next') or request.POST.get('next') or 'panel_productos'
    form = ProductoForm(request.POST or None, request.FILES or None, instance=producto)
    
    if request.method == 'POST' and form.is_valid():
        prod_editado = form.save()
        cambios = []

        # Eliminar imágenes seleccionadas para borrar
        eliminar_ids = request.POST.getlist('eliminar_imagenes')
        if eliminar_ids:
            borradas_count = ImagenProducto.objects.filter(producto=prod_editado, id__in=eliminar_ids).delete()[0]
            if borradas_count > 0:
                cambios.append(f"{borradas_count} imágenes de galería eliminadas")

        # Subir nuevas imágenes adicionales
        imagenes_extra = request.FILES.getlist('imagenes_adicionales')
        if imagenes_extra:
            orden_base = ImagenProducto.objects.filter(producto=prod_editado).count()
            for idx, file in enumerate(imagenes_extra):
                ImagenProducto.objects.create(
                    producto=prod_editado,
                    imagen=file,
                    orden=orden_base + idx
                )
            cambios.append(f"{len(imagenes_extra)} nuevas imágenes agregadas a la galería")

        if nombre_ant != prod_editado.nombre:
            cambios.append(f"Nombre: '{nombre_ant}' ➔ '{prod_editado.nombre}'")
        if precio_ant != prod_editado.precio:
            cambios.append(f"Precio: ${precio_ant} ➔ ${prod_editado.precio}")
        if stock_ant != prod_editado.stock:
            cambios.append(f"Stock: {stock_ant} ➔ {prod_editado.stock}")
        if dispon_ant != prod_editado.disponible:
            cambios.append(f"Disponible: {'Sí' if dispon_ant else 'No'} ➔ {'Sí' if prod_editado.disponible else 'No'}")
        
        cat_nueva = prod_editado.categoria.nombre if prod_editado.categoria else "Sin categoría"
        if cat_ant != cat_nueva:
            cambios.append(f"Categoría: '{cat_ant}' ➔ '{cat_nueva}'")

        if 'imagen' in request.FILES:
            cambios.append("Nueva imagen de portada subida")

        detalles_str = " | ".join(cambios) if cambios else "Edición realizada sin cambios principales"
        usuario_log = request.user if request.user.is_authenticated else None

        LogProducto.objects.create(
            producto_id=prod_editado.id,
            nombre_producto=prod_editado.nombre,
            accion='EDICION',
            usuario=usuario_log,
            detalles=detalles_str
        )

        messages.success(request, f"Producto '{prod_editado.nombre}' actualizado.")
        return redirect(next_url)
        
    return render(request, "panel/producto_form.html", {
        "form": form,
        "editando": True,
        "next": next_url
    })


@staff_member_required(login_url='login')
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    usuario_log = request.user if request.user.is_authenticated else None
    detalles_str = f"Producto eliminado definitivamente. Nombre final: '{producto.nombre}', Precio final: ${producto.precio}, Stock final: {producto.stock}, Categoría: {producto.categoria.nombre if producto.categoria else 'Sin categoría'}"
    
    LogProducto.objects.create(
        producto_id=producto.id,
        nombre_producto=producto.nombre,
        accion='ELIMINACION',
        usuario=usuario_log,
        detalles=detalles_str
    )
    
    producto.delete()
    messages.success(request, f"Producto '{producto.nombre}' eliminado.")
    return redirect('panel_productos')


@staff_member_required(login_url='login')
def panel_pedidos(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')

    pedidos = Pedido.objects.all()

    if q:
        pedidos = pedidos.filter(
            models.Q(nombre_completo__icontains=q) |
            models.Q(rut__icontains=q) |
            models.Q(email__icontains=q) |
            models.Q(id__icontains=q.replace('#', ''))
        )

    if estado:
        pedidos = pedidos.filter(estado=estado)

    pedidos = pedidos.order_by('-id')

    return render(request, 'panel/pedidos.html', {
        'pedidos': pedidos,
        'q': q,
        'estado_filtro': estado,
        'estados_choices': Pedido.ESTADO_CHOICES
    })


@staff_member_required(login_url='login')
def detalle_pedido(request, id):
    pedido = get_object_or_404(Pedido.objects.prefetch_related('items__producto'), id=id)
    return render(request, 'panel/detalle_pedido.html', {'pedido': pedido, 'estados_choices': Pedido.ESTADO_CHOICES})


@staff_member_required(login_url='login')
def actualizar_estado_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        empresa = request.POST.get('empresa_transporte', '').strip()
        seguimiento = request.POST.get('numero_seguimiento', '').strip()

        if nuevo_estado in dict(Pedido.ESTADO_CHOICES):
            pedido.estado = nuevo_estado
        pedido.empresa_transporte = empresa
        pedido.numero_seguimiento = seguimiento
        
        if nuevo_estado == 'PAGADO' and not pedido.pagado:
            pedido.confirmar_pago()
        else:
            pedido.save()

        usuario_log = request.user if request.user.is_authenticated else None
        detalles_act = f"Estado cambiado a '{pedido.get_estado_display()}'"
        if empresa or seguimiento:
            detalles_act += f" | Transporte: '{empresa if empresa else 'No especificado'}' | Seguimiento: '{seguimiento if seguimiento else 'Sin código'}'"

        LogPedido.objects.create(
            pedido_id=pedido.id,
            codigo_orden=pedido.codigo_orden,
            cliente_nombre=pedido.nombre_completo,
            cliente_email=pedido.email,
            accion='ESTADO_CAMBIO',
            usuario=usuario_log,
            detalles=detalles_act
        )

        messages.success(request, f"Estado del pedido #{pedido.codigo_orden} actualizado a '{pedido.get_estado_display()}'.")
    return redirect('detalle_pedido', id=id)


@staff_member_required(login_url='login')
def enviar_seguimiento_email(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if not pedido.numero_seguimiento:
        messages.error(request, "Debes ingresar un número de seguimiento primero.")
        return redirect('detalle_pedido', id=id)

    try:
        asunto = f"📦 Tu pedido #{pedido.codigo_orden} de Rara Tienda ya va en camino"
        mensaje = f"""Hola {pedido.nombre_completo},

¡Te tenemos excelentes noticias! Tu pedido #{pedido.codigo_orden} ha sido enviado.

Detalles del despacho:
- Transporte: {pedido.empresa_transporte or 'Empresa de Envíos'}
- Código / Nº de Seguimiento: {pedido.numero_seguimiento}

Puedes realizar el seguimiento de tu paquete directamente en el sitio web del transporte.

¡Muchas gracias por comprar en Rara Tienda! 🦉
"""
        send_mail(
            asunto,
            mensaje,
            settings.DEFAULT_FROM_EMAIL,
            [pedido.email],
            fail_silently=False
        )

        usuario_log = request.user if request.user.is_authenticated else None
        LogPedido.objects.create(
            pedido_id=pedido.id,
            codigo_orden=pedido.codigo_orden,
            cliente_nombre=pedido.nombre_completo,
            cliente_email=pedido.email,
            accion='SEGUIMIENTO',
            usuario=usuario_log,
            detalles=f"Correo de despacho enviado a {pedido.email} | Transporte: {pedido.empresa_transporte or 'No especificado'} | Nº Seguimiento: {pedido.numero_seguimiento}"
        )

        messages.success(request, f"Correo con número de seguimiento enviado a {pedido.email}.")
    except Exception as e:
        messages.error(request, f"No se pudo enviar el correo: {e}")

    return redirect('detalle_pedido', id=id)


@staff_member_required(login_url='login')
def enviar_resena_email(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    url_resena = pedido.get_enlace_resena()
    
    asunto = f"¡Cuéntanos tu experiencia con el Pedido #{pedido.codigo_orden}! 🌟 — Rara Tienda"
    mensaje = f"""Hola {pedido.nombre_completo},

¡Esperamos que estés disfrutando tus productos de Rara Tienda!

Nos encantaría saber tu opinión sobre tu compra. Tu valoración ayuda a otros amantes de la naturaleza a conocer nuestros productos:

👉 Evalúa tus productos aquí: {url_resena}

¡Muchas gracias por apoyar nuestra tienda!

Con cariño,
El equipo de Rara Tienda
https://www.raratienda.cl/
"""
    try:
        send_mail(
            asunto,
            mensaje,
            settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'contacto@raratienda.cl',
            [pedido.email],
            fail_silently=False
        )
        messages.success(request, f"Correo de reseña enviado con éxito a {pedido.email}.")
    except Exception as e:
        messages.error(request, f"Error al enviar el correo: {e}")
        
    return redirect('detalle_pedido', id=pedido.id)


@staff_member_required(login_url='login')
def confirmar_pago_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    if request.method == 'POST':
        with transaction.atomic():
            pedido = Pedido.objects.select_for_update().get(id=id)
            if pedido.pagado:
                messages.info(request, f'El pedido #{pedido.codigo_orden} ya estaba pagado.')
                return redirect('detalle_pedido', id=pedido.id)
            
            pedido.confirmar_pago()
            pedido.estado = 'PAGADO'
            pedido.save()

            usuario_log = request.user if request.user.is_authenticated else None
            LogPedido.objects.create(
                pedido_id=pedido.id,
                codigo_orden=pedido.codigo_orden,
                cliente_nombre=pedido.nombre_completo,
                cliente_email=pedido.email,
                accion='PAGO_OK',
                usuario=usuario_log,
                detalles="Pago confirmado manualmente por administrador desde el panel"
            )
        
        asunto = f'¡Pago Confirmado! Pedido #{pedido.codigo_orden} en Rara Tienda 🦉'
        mensaje = f'''Hola {pedido.nombre_completo},

¡Tenemos excelentes noticias! Hemos confirmado el pago de tu pedido.

Tu nido de productos ya está siendo preparado con mucho cariño para ser enviado a {pedido.direccion}, {pedido.ciudad}. 

Te avisaremos por esta misma vía en cuanto el paquete inicie su vuelo.

¡Gracias por apoyar el arte y la naturaleza!
El equipo de Rara Tienda.
'''
        try:
            send_mail(
                asunto, 
                mensaje, 
                settings.DEFAULT_FROM_EMAIL, 
                [pedido.email], 
                fail_silently=False
            )
            messages.success(request, f'Pago confirmado, stock actualizado y correo enviado a {pedido.email}.')
        except Exception as e:
            messages.warning(request, f'Pago confirmado y stock descontado, pero no se pudo enviar el correo automático.')
            print(f"Error SMTP: {e}")
            
    return redirect('detalle_pedido', id=pedido.id)


@staff_member_required(login_url='login')
def panel_cupones(request):
    cupones = Cupon.objects.all().order_by('-id')
    return render(request, 'panel/cupones.html', {'cupones': cupones})


@staff_member_required(login_url='login')
def crear_cupon(request):
    form = CuponForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Cupón creado exitosamente.")
        return redirect('panel_cupones')
    return render(request, 'panel/cupon_form.html', {'form': form, 'editando': False})


@staff_member_required(login_url='login')
def editar_cupon(request, id):
    cupon = get_object_or_404(Cupon, id=id)
    form = CuponForm(request.POST or None, instance=cupon)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f"Cupón '{cupon.codigo}' actualizado exitosamente.")
        return redirect('panel_cupones')
    return render(request, 'panel/cupon_form.html', {'form': form, 'editando': True})


@staff_member_required(login_url='login')
def toggle_cupon(request, id):
    cupon = get_object_or_404(Cupon, id=id)
    cupon.activo = not cupon.activo
    cupon.save()
    messages.success(request, f"Estado del cupón '{cupon.codigo}' cambiado.")
    return redirect('panel_cupones')


@staff_member_required(login_url='login')
def eliminar_cupon(request, id):
    cupon = get_object_or_404(Cupon, id=id)
    if request.method == 'POST':
        cupon.delete()
        messages.success(request, "Cupón eliminado.")
    return redirect('panel_cupones')


@staff_member_required(login_url='login')
def exportar_pedidos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Rara Tienda"

    fill_header = PatternFill(start_color="FCA311", end_color="FCA311", fill_type="solid")
    font_header = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(left=Side(style='thin', color='DDDDDD'),
                         right=Side(style='thin', color='DDDDDD'),
                         top=Side(style='thin', color='DDDDDD'),
                         bottom=Side(style='thin', color='DDDDDD'))

    headers = ['ID Orden', 'Cliente', 'RUT', 'Email', 'Teléfono', 'Ciudad', 'Dirección', 'Estado', 'Pagado', 'Total', 'Transporte', 'Nº Seguimiento', 'Fecha']
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 18

    pedidos = Pedido.objects.all().order_by('-id')
    
    for row_num, p in enumerate(pedidos, start=2):
        ws.append([
            f"#{p.codigo_orden}",
            p.nombre_completo,
            p.rut,
            p.email,
            f"+56{p.telefono}",
            p.ciudad,
            p.direccion,
            p.get_estado_display(),
            "Sí" if p.pagado else "No",
            f"${p.get_total_cost() - p.descuento_aplicado}",
            p.empresa_transporte,
            p.numero_seguimiento,
            p.creado.strftime("%Y-%m-%d %H:%M")
        ])
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [1, 3, 5, 8, 9, 10, 13]:
                cell.alignment = align_center

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Pedidos_Rara_Tienda.xlsx"'
    wb.save(response)
    return response


@staff_member_required(login_url='login')
def exportar_stock_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Rara Tienda"

    fill_header = PatternFill(start_color="FCA311", end_color="FCA311", fill_type="solid")
    font_header = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(left=Side(style='thin', color='DDDDDD'),
                         right=Side(style='thin', color='DDDDDD'),
                         top=Side(style='thin', color='DDDDDD'),
                         bottom=Side(style='thin', color='DDDDDD'))

    font_agotado = Font(color="E74C3C", bold=True)
    font_disponible = Font(color="27AE60", bold=True)

    headers = ['ID', 'Producto', 'Stock Actual', 'Precio', 'Estado']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

    productos = Producto.objects.all().order_by('nombre')
    
    for row_num, p in enumerate(productos, start=2): 
        estado = "Disponible" if p.stock > 0 else "Agotado"
        ws.append([p.id, p.nombre, p.stock, f"${p.precio}", estado])
        
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [1, 3, 4, 5]: 
                cell.alignment = align_center
            if col_num == 5:
                cell.font = font_agotado if estado == "Agotado" else font_disponible

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventario_Rara_Tienda.xlsx"'
    wb.save(response)
    return response


@staff_member_required(login_url='login')
def panel_configuracion(request):
    config = ConfiguracionSitio.get_solo()
    if request.method == 'POST':
        form = ConfiguracionSitioForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración del sitio actualizada correctamente.")
            return redirect('panel_configuracion')
    else:
        form = ConfiguracionSitioForm(instance=config)
    return render(request, "panel/configuracion.html", {"form": form, "config": config})


@staff_member_required(login_url='login')
def panel_blog(request):
    posts = BlogPost.objects.all()
    return render(request, "panel/blog_list.html", {"posts": posts})


@staff_member_required(login_url='login')
def crear_blog_post(request):
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Artículo publicado exitosamente.")
            return redirect('panel_blog')
    else:
        form = BlogPostForm()
    return render(request, "panel/blog_form.html", {"form": form, "titulo_pagina": "Nuevo Artículo del Blog"})


@staff_member_required(login_url='login')
def editar_blog_post(request, id):
    post = get_object_or_404(BlogPost, id=id)
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            form.save()
            messages.success(request, "Artículo actualizado correctamente.")
            return redirect('panel_blog')
    else:
        form = BlogPostForm(instance=post)
    return render(request, "panel/blog_form.html", {"form": form, "post": post, "titulo_pagina": "Editar Artículo"})


@staff_member_required(login_url='login')
def eliminar_blog_post(request, id):
    post = get_object_or_404(BlogPost, id=id)
    if request.method == 'POST':
        post.delete()
        messages.success(request, "Artículo eliminado.")
    return redirect('panel_blog')


@staff_member_required(login_url='login')
def panel_resenas(request):
    resenas = ResenaProducto.objects.select_related('producto').all()
    return render(request, "panel/resenas.html", {"resenas": resenas})


@staff_member_required(login_url='login')
def toggle_resena(request, id):
    resena = get_object_or_404(ResenaProducto, id=id)
    resena.aprobado = not resena.aprobado
    resena.save()
    estado = "aprobada" if resena.aprobado else "ocultada"
    messages.info(request, f"Reseña {estado}.")
    return redirect('panel_resenas')


@staff_member_required(login_url='login')
def eliminar_resena(request, id):
    resena = get_object_or_404(ResenaProducto, id=id)
    if request.method == 'POST':
        resena.delete()
        messages.success(request, "Reseña eliminada.")
    return redirect('panel_resenas')


@staff_member_required(login_url='login')
def panel_guia(request):
    return render(request, "panel/guia.html")
